# libraries
import openpyxl
import os
from . import settings

# classes
from openpyxl.cell.cell import Cell
from openpyxl.styles.fonts import Font
from typing import List
from .Notebook import Notebook
from openpyxl.worksheet.worksheet import Worksheet

def create_workbook() -> openpyxl.Workbook:
    """
    Creates a workbook with an audit sheet for an Excel file
    """
    wb = openpyxl.Workbook()
    wb.create_sheet('Audit')
    try: wb.remove(wb['Sheet'])
    except: pass
    ws_xlsx = wb['Audit']
    ws_xlsx.cell(1,1,'Run By:')
    ws_xlsx.cell(1,2,str(os.environ.get('USERNAME')))
    ws_xlsx.cell(2,1,'DB_Checker Version:')
    ws_xlsx.cell(2,2,settings.version)
    ws_xlsx.cell(3,1,'Workspace path:')
    ws_xlsx.cell(3,2,settings.workspace_path)
    ws_xlsx.cell(4,1,'Force same extension:')
    ws_xlsx.cell(4,2,str(settings.check_ext))
    ws_xlsx.cell(5,1,'Scrap identifiers:')
    ws_xlsx.cell(5,2,', '.join(settings.scrap_contains + settings.scrap_startswith + settings.scrap_endswith))
    return wb

def fit_columns(wb: openpyxl.Workbook) -> None:
    """
    For all sheets in a workbook, adjust column length to fit all contents.

    Args:
        wb (openpyxl.Workbook): The workbook object to edit
    """
    for sheets in wb.worksheets:
        for col in sheets.columns:
            max_length = 0
            column = col[0].column_letter # pyright: ignore[reportUnknownVariableType, reportAttributeAccessIssue, reportUnknownMemberType]
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.2
            sheets.column_dimensions[column].width = adjusted_width

def hyperlink(cell: Cell, hyperlink: str) -> None:
    """
    Adds a hyperlink and standard formatting to a cell

    Args:
        cell (openpyxl.Cell):   The is the cell that will receive the link and formatting.
        hyperlink (str):        This is the hyperlink that will be added to the cell.
    """
    cell.hyperlink = hyperlink
    cell.font = Font(color='0000FF', underline='single')

def write_headers(ws: Worksheet) -> None:
    """
    Writes default headers for a new sheet

    Args:
        ws (openpyxl.Worksheet):    This is the worksheet that will have the new headers.
    """
    headers = [
        'Support',
        'Notebook Name',
        'QRM Status',
        'Similarity',
        'Signatures',
        'Downloaded by DB Checker',
        'Source File',
        'Notebook URL'
    ]
    for i in range(1, len(headers) + 1): ws.cell(1,i,headers[i-1])

def write_data(ws: Worksheet, notebooks: List[Notebook]) -> None:
    """
    Writes notebooks' data into a worksheet

    Args:
        ws (openpyxl.Worksheet):    This is the worksheet that will have the new data.
        notebooks (List[Notebook]): List of notebooks whose information will be written
    """
    # loop through cells, skip headers
    for i in range(2, len(notebooks)+2):
        nb = notebooks[i-2]

        if nb.qrm: qrm = 'OK.'
        elif nb.source_path is None: qrm = 'MISSING'
        if nb.qrm is None: qrm = 'REVIEW REQUIRED'
        else: qrm = 'ISSUE'

        if not settings.check_similarity and not settings.levenshtein: s = 'N/A'
        elif nb.source_path is None: s = 'N/A'
        elif nb.local is None or nb.similarity is None: s = 'Failed to comapre file'
        elif settings.levenshtein: s = str(round(nb.similarity * 100,2)) + '%'
        else:
            if bool(nb.similarity): s = 'Exact match'
            else: s = 'Not an exact match'
        
        if not settings.download: d = 'N/A'
        elif nb.downloaded is None: d = 'N/A'
        elif nb.downloaded: d = 'Yes'
        else: d = 'Failed'

        if nb.source_path is None: source_path = ''
        else: source_path = nb.source_path

        data: List[str] = [
            nb.support,
            nb.name,
            qrm,
            s,
            nb.signatures,
            d,
            source_path,
            nb.url
        ]

        for j in range(1, len(data)): ws.cell(i,j,data[j-1])

        url_cell = ws.cell(i,len(data),data[-1])
        hyperlink(url_cell, nb.url)

def save(wb: openpyxl.Workbook, path: str, confirmation: bool = True) -> None:
    """
    Attempt to save an Excel workbook, allowing retries if the file is open

    Args:
        wb (openpyxl.workbook): openpyxl workbook object
        path (str):             Path of Excel workbook
        confirmation (bool):    print confirmation message
    """
    saved = False
    while not saved:
        try:
            wb.save(path)
            wb.close()
            if confirmation: print(f'Saved {path}')
            saved = True
        except PermissionError:
            print('Failed to save workbook. Please ensure that it is closed before continuing.')
            inp = input("Type SKIP to continue or press [ENTER] to try again...")
            if inp.lower().strip() == 'skip':
                wb.close()
                saved = True
    pass